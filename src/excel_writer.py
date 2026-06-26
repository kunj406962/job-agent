from pathlib import Path
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

COLUMNS = [
    "id", "date_found", "date_posted", "company", "position",
    "location", "apply_url", "source", "description_snippet",
    "deadline", "status", "notes",
]

XLSX_PATH = Path(__file__).parent.parent / "data" / "jobs.xlsx"
ARCHIVE_PATH = Path(__file__).parent.parent / "data" / "jobs_archive.xlsx"

ARCHIVE_THRESHOLD = 500  # rows before archiving old entries

# Colours
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")   # new rows
GREY_FILL  = PatternFill(fill_type="solid", fgColor="D9D9D9")   # old rows
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")

COL_WIDTHS = {
    "A": 20, "B": 12, "C": 12, "D": 20, "E": 35,
    "F": 25, "G": 50, "H": 18, "I": 50, "J": 12, "K": 14, "L": 30,
}


def load_existing_ids() -> set:
    """Return set of all job IDs already in jobs.xlsx."""
    if not XLSX_PATH.exists():
        return set()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]] # type: ignore
    if "id" not in headers:
        return set()

    id_col = headers.index("id") + 1
    return {
        row[0]
        for row in ws.iter_rows(min_row=2, min_col=id_col, max_col=id_col, values_only=True) # type: ignore
        if row[0]
    }


def write_jobs(new_jobs: list) -> int:
    """
    1. Load existing jobs.xlsx (or create it).
    2. Archive grey rows if over threshold.
    3. Turn all currently green rows grey.
    4. Append new_jobs as green rows.
    5. Save.
    Returns number of new rows written.
    """
    if not new_jobs:
        return 0

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if XLSX_PATH.exists():
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs" # type: ignore
        _write_header(ws)

    # Archive if over threshold
    data_rows = ws.max_row - 1  # type: ignore # exclude header
    if data_rows >= ARCHIVE_THRESHOLD:
        _archive_grey_rows(ws)

    # Turn existing green rows grey
    _age_green_rows(ws)

    # Append new rows in green
    for job in new_jobs:
        row = [
            job.get("id", ""),
            today,
            job.get("date_posted", ""),
            job.get("company", ""),
            job.get("position", ""),
            job.get("location", ""),
            job.get("apply_url", ""),
            job.get("source", ""),
            job.get("description_snippet", ""),
            job.get("deadline", ""),
            "Not Applied",
            "",
        ]
        ws.append(row) # type: ignore
        # Colour the new row green
        for cell in ws[ws.max_row]: # type: ignore
            cell.fill = GREEN_FILL

    wb.save(XLSX_PATH)
    return len(new_jobs)


def _age_green_rows(ws):
    """Turn any green rows grey (they were new last run, now they're old)."""
    for row in ws.iter_rows(min_row=2):
        if row[0].fill and row[0].fill.fgColor.rgb == GREEN_FILL.fgColor.rgb:
            for cell in row:
                cell.fill = GREY_FILL


def _archive_grey_rows(ws):
    """Move all grey rows into jobs_archive.xlsx and delete them from ws."""
    grey_rows = []
    rows_to_delete = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].fill and row[0].fill.fgColor.rgb == GREY_FILL.fgColor.rgb:
            grey_rows.append([cell.value for cell in row])
            rows_to_delete.append(row_idx)

    if not grey_rows:
        return

    # Write to archive
    if ARCHIVE_PATH.exists():
        awb = openpyxl.load_workbook(ARCHIVE_PATH)
        aws = awb.active
    else:
        awb = openpyxl.Workbook()
        aws = awb.active
        aws.title = "Archive" # type: ignore
        _write_header(aws)

    for row_data in grey_rows:
        aws.append(row_data) # type: ignore

    awb.save(ARCHIVE_PATH)
    print(f"[excel] Archived {len(grey_rows)} old rows to jobs_archive.xlsx")

    # Delete grey rows from active sheet (reverse order to preserve indices)
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)


def _write_header(ws):
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"